import pandas as pd
import numpy as np
from pandas import ExcelWriter
from scipy.stats import expon
import rdfpandas,xlsxwriter,string, os
from ..ResnetAPI.NetworkxObjects import PSObject
from openpyxl import load_workbook
from pandas.api.types import is_string_dtype,is_numeric_dtype


MIN_COLUMN_WIDTH = 0.65 # in inches
NUMBER_OF_REFERENCE = 'Number of references'
MAX_TAB_LENGTH = 31

class df(pd.DataFrame):
  # re-writing parent pd.DataFrame function is a BAD idea. Use other names
  pass
  header_format = {
                  'bold': True,
                  'text_wrap': True,
                  'valign': 'vjustify',#'bottom',#'top', #vcenter',
                  'align': 'center', # 'left'
                  'height':15
                  #'fg_color': '#D7E4BC',
                  #'border': 1,
                  }
  # need to declare format dicts explicitly to avoid pd.DataFrame warnings in stdr
  column2format = dict() 
  conditional_frmt = dict()
  tab_format = dict()
  col2rank = dict()
  _name_ = ''

  def __init__(self, *args, **kwargs):
      '''
      kwargs:
          name:str
      '''
      dfname = kwargs.pop('name','')  
      pd.DataFrame.__init__(self,*args, **kwargs)
      self._name_ = dfname
      self.column2format = dict() # {column_index:{'font_color':'blue'}}
      self.col2rank = dict() # {colname:rank(int)}, used for column ranking and sorting by SemanticSearch
      self.conditional_frmt = dict() # {area:{conditional_format}}
      self.tab_format = dict() # font and color of the Excel tab with worksheet


  @classmethod 
  def from_pd(cls,d:pd.DataFrame,dfname=''):
      new_df = cls(d)
      new_df._name_ = dfname
      return new_df


  def copy_format(self, from_df:'df'):
      self.header_format = from_df.header_format.copy()
      self.column2format = from_df.column2format.copy()
      self.conditional_frmt = from_df.conditional_frmt.copy()
      self.tab_format = from_df.tab_format.copy()
      self.col2rank = from_df.col2rank.copy()


  def __copy_attrs(self,from_df:'df'):
      self._name_ = from_df._name_
      self.copy_format(from_df)


  def max_colrank(self):
    return max(self.col2rank.values()) if self.col2rank else 0


  def dfcopy(self,only_columns:list=[], rename2:dict=dict(),deep=True) ->'df':
      '''
      input:
        rename2 - map of column names to rename {old_name:new_name}
      '''
      # re-writing parent pd.DataFrame function is a BAD idea. Use other names
      newpd = super().copy(deep)
      missing_columns = list()
      my_columns = list()
      if only_columns:
        [(missing_columns,my_columns)[c in self.columns].append(c) for c in only_columns]
        if len(my_columns) < len(only_columns):
          print(f'Worksheet {self._name_} is missing {missing_columns} columns')
        newpd = newpd[my_columns]

      if rename2:
        newpd = newpd.rename(columns=rename2)
      newpd = newpd.reindex()

      newdf = df.from_pd(newpd,self._name_)
      newdf.copy_format(self)
      column2format = newdf.column2format
      col2rank = newdf.col2rank
      for old_name, new_name in rename2.items():
        if old_name in newdf.column2format:
          column2format[new_name] = column2format.get(old_name)
        if old_name in newdf.col2rank:
          col2rank[new_name] = column2format.get(old_name)
      return newdf


  @staticmethod
  def _hyperlink(identifier:str,url:str,display_str=''):
      # hyperlink in Excel does not work with long URLs, 
      display_str = display_str if display_str else identifier
      return '=HYPERLINK("'+url+identifier+'",\"{}\")'.format(display_str)
  

  def add_format(self, from_df:'df'):
    '''
    updates formats with from_df format
    '''
    self.header_format.update(from_df.header_format)
    self.column2format.update(from_df.column2format)
    self.conditional_frmt.update(from_df.conditional_frmt)
    self.tab_format.update(from_df.tab_format)
    self.col2rank.update({col:max(rank,self.col2rank.get(col,rank)) for col,rank in from_df.col2rank.items()})


  def set_col_format(self,fmt:dict):
      self.column2format = fmt


  def add_column_format(self,column_name:str,fmt_property:str,fmt_value):
      try:
          self.column2format[column_name].update({fmt_property:fmt_value})
      except KeyError:
          self.column2format[column_name] = {fmt_property:fmt_value}


  def set_hyperlink_color(self, column_names:list):
      [self.add_column_format(column_name,'font_color','blue') for column_name in column_names]


  def inch_width(self,col:str|int):
      my_column = col if isinstance(col,str) else str(self.columns[col])
      try:
          return self.column2format[my_column]['inch_width']
      except KeyError:
          return 0


  def set_inch_width(self,layout:dict):
      """
      layout = {column_index:inch_width}
      """
      for col_idx, width in layout.items():
          col_name = str(self.columns[col_idx])
          self.column2format[col_name] = {'inch_width':width}


  def inch_width_layout(self):
      layout = {i:self.inch_width(i) for i in range(0,len(self.columns))}
      return layout


  @classmethod
  def read(cls, *args, **kwargs):
      '''
      Input
      -----
      args[0] - input filname for reading Excel file add sheet_name=my_worksheet_name to kwargs
      kwargs = {sheet_name:str, read_formula:bool,names=[]}
      '''
      df_name = kwargs.pop('name','')
      fname = str(args[0])
      extension = os.path.splitext(fname)[1][1:] # remove period
      if extension == 'xlsx':
          read_formula = kwargs.pop('read_formula',False)
          if read_formula:
              wb = load_workbook(filename=fname)
              try:
                  sheet = wb[kwargs['sheet_name']]  
              except KeyError:
                  #sheet_names = wb.get_sheet_names()
                  sheet = wb[str(wb.sheetnames[0])]

              header_pos = kwargs.pop('header',0)
              skiprows = kwargs.pop('skiprows',0)
              _dfname_ = df_name if df_name else sheet
              _pd = pd.DataFrame(list(sheet.values))
              _pd.columns = _pd.iloc[header_pos].to_list()
              _df = df.from_pd(_pd[header_pos+1+skiprows:],str(_dfname_))
              return _df
          try:
              _pd = pd.read_excel(*args, **kwargs)
              _df = df.from_pd(_pd,df_name)
              # _df._name_ = df_name
              return _df
          except FileNotFoundError: 
              raise(FileNotFoundError)
              #return df()
      elif extension in ('tsv','txt','tab'):
          kwargs.pop('sheet_name','')
          kwargs.pop('read_formula',False)
          my_kwargs = dict(kwargs)
          my_kwargs['sep'] = '\t'
          try:
              return df.from_pd(pd.DataFrame(pd.read_csv(*args,**my_kwargs)),df_name)
          except FileNotFoundError:
              raise(FileNotFoundError)
              #return df()
      elif extension == ('csv'):
          kwargs.pop('sheet_name','')
          kwargs.pop('read_formula',False)
          try:
              my_kwargs = dict(kwargs)
              my_kwargs['sep'] = ','
              _df =  df.from_pd(pd.DataFrame(pd.read_csv(*args,**my_kwargs)),df_name)
              return _df
          except FileNotFoundError:
              raise(FileNotFoundError)
              #return df()
      return df()


  def pandas2rdf(self, remap:dict):
      rdf_pandas = pd.DataFrame(self)
      for c in rdf_pandas.columns:
          rdf_pandas.rename(columns=remap[c])
          return rdfpandas.to_graph(rdf_pandas)


  def apply_and_concat(self, field, func, column_names):
          merged_pd = pd.concat((self,self[field].apply(lambda cell: pd.Series(func(field,cell),index=column_names))),axis=1)
          to_return = df.from_pd(merged_pd)
          to_return.__copy_attrs(self)
          return to_return

  #not yet tested
  def apply_and_concat2(self, concept_name, func, column_names):
          return pd.concat((self,self['Name'].apply(lambda cell: pd.Series(func(cell,concept_name),index=column_names))),axis=1)


  @classmethod 
  def from_dict(cls, dic:dict, **kwargs):
      '''
      Input
      -----
      dic - {col_name:[values]}
      orient: Literal['columns', 'index', 'tight'] = ...,
      dtype: _str = ...,
      columns: list[_str] = ...\n
      The "orientation" of the data. If the keys of the passed dict should be the columns of the resulting DataFrame, pass 'columns' (default). Otherwise if the keys should be rows, pass 'index'. If 'tight', assume a dict with keys ['index', 'columns', 'data', 'index_names', 'column_names']
      '''
      df_name = kwargs.pop('name','')
      new_pd = cls.from_pd(pd.DataFrame.from_dict(dic,**kwargs))
      new_pd._name_ = df_name
      return new_pd


  @classmethod 
  def from_rows(cls,rows:list[list],header:list[str],index=-1,dfname=''):
      '''
      rows - list/set of lists/tuples
      '''
      _2return = pd.DataFrame(rows,columns=header)
      if index >= 0:
          _2return.set_index(header[index], inplace=True)
      return df.from_pd(_2return,dfname)
  

  @classmethod
  def from_dict2(cls,dic:dict, key_colname:str, value_colname:str)->'df':
      #new_df = cls(pd.DataFrame(columns=[key_colname,value_colname]))
      new_df = df.from_pd(pd.DataFrame.from_dict({key_colname:list(dic.keys()), value_colname:list(dic.values())}))
      return new_df
  

  def merge_dict(self, dict2add:dict[str,str|int|float|tuple], new_col:str, map2column:str, 
                 add_all=False, case_sensitive_match=False,default_val = ''):
      '''
      input:
        "dict2add" must have keys equal to values in "map2column"
      output:
        new df with values from dict2add.values() in "new_col" 
      '''
      in2df = self.dfcopy()
      mapping_column = map2column
      if case_sensitive_match:
        my_dict = dict2add
      else:
        my_dict = {k.lower():v for k,v in dict2add.items()}
        mapping_column = map2column+'lowercase'
        in2df[mapping_column] = in2df[map2column].str.lower()
    
      pd2merge = df.from_dict2(my_dict,mapping_column,new_col)
      how = 'outer' if add_all else 'left'
      merged_df = in2df.merge_df(pd2merge,how=how, on=mapping_column)
      merged_df[new_col] = merged_df[new_col].fillna(default_val)
      if not case_sensitive_match:
        my_cols = self.columns.to_list() + [new_col]
        merged_df = merged_df.dfcopy(my_cols)
      return merged_df


  def append_df(self, other:'df'):
    '''
    output:
      reindexed concatinated df
    '''
    merged_pd = pd.concat([self,other],ignore_index=True)
    merged_pd = merged_pd.reindex()
    merged_df = df.from_pd(merged_pd,dfname=self._name_)
    merged_df.copy_format(other)
    merged_df.add_format(from_df=self)
    return merged_df
  

  def add_dict(self, dic:dict):
    new_df = df([dic])
    return self.append_df(new_df)
  

  @staticmethod
  def concat_df(dfs:list, dfname=''):
      merged_pd = pd.concat(dfs,ignore_index=True)
      merged_pd = merged_pd.reindex()
      merged_df = df.from_pd(merged_pd,dfname=dfname)
      return merged_df
  
  
  def reindex_df(self):
      reindexed_pd = self.reindex()
      reindexed_df = df.from_pd(reindexed_pd,dfname=self._name_)
      reindexed_df.copy_format(self)
      return reindexed_df


  def merge_df(self, df2merge:'df', **kwargs)->'df':
      '''
      Input
      -----
      right df = args[0]\n
      non pd.DataFrame kwargs: 'name','columns'\n
      "on" kwarg has to specify column in both self and right df
      "how" - {"left", "right", "outer", "inner", "cross"}, default "inner"
      inner: use intersection of keys from both frames,preserve the order of the left keys.

      Return
      ------
      df merged using "on" kwarg with columns added from right df specified by 'columns'\n
      if 'columns' were not specified will merge all columns from right df\n
      Format of "self" takes precedent
      '''
      df_name = kwargs.pop('name',self._name_)
      columns2copy = list(kwargs.pop('columns',[]))
      merge_on_column = kwargs.get('on')
      if columns2copy:
          columns2copy.append(merge_on_column)
          my_df2merge = df2merge.dfcopy(columns2copy)
      else:
          my_df2merge = df2merge

      merged_df = df.from_pd(self.merge(pd.DataFrame(my_df2merge), **kwargs),dfname=df_name)
      merged_df.add_format(from_df=df2merge)
      merged_df.add_format(from_df=self)
      return merged_df


  @classmethod
  def psobj2df(cls,obj:PSObject,key_colname:str,value_colname:str,dfname=str()):
      key_col = list()
      value_col = list()
      for k,v_list in obj.items():
          key_col += [k]*len(v_list)
          value_col += v_list
      
      new_df = cls.from_dict({key_colname:key_col,value_colname:value_col})
      new_df._name_= dfname
      return new_df


  def merge_psobject(self, obj:PSObject, new_col:str, map2column:str, 
          add_all=False, values21cell=False, sep=';', case_sensitive_match = False):

      in2pd = self.dfcopy()
      if not case_sensitive_match:
          in2pd[map2column] = self[map2column].apply(lambda x: str(x).lower())

      if values21cell:
          if case_sensitive_match:
              merge_dict = {k:sep.join(v) for k,v in obj.items()}
          else:
              merge_dict = {str(k).lower():sep.join(v) for k,v in obj.items()}
              
          in2pd =  in2pd.merge_dict(merge_dict,new_col,map2column,add_all)
          in2pd[map2column] = self[map2column]
      else:          
          obj_df = df.psobj2df(obj,map2column,new_col)
          how = 'outer' if add_all else 'left'
          in2pd = in2pd.merge_df(obj_df,how=how,on=map2column)
      
      in2pd.__copy_attrs(self)
      return in2pd


  def get_rows(self, by_value1, in_column1, and_by_value2, in_column2):
      return self.loc[(self[in_column1] == by_value1) & (self[in_column2] == and_by_value2)]


  def get_cells(self,by_value1, in_column1, and_by_value2, in_column2, from_column3):
      return (self.loc[(self[in_column1] == by_value1) & (self[in_column2] == and_by_value2),from_column3]).iloc[0]


  def df2json(self, to_file:str, dir=''):
      dump_fname = to_file
      if  to_file[-5:] != '.json': dump_fname += '.json'
      dump_fname = dir+dump_fname
      with open(dump_fname, 'w') as jsonout:
          self.reset_index(inplace=True)
          jsonout.write(self.to_json(None,indent=2, orient='index'))
  

  def __worksheet_area4(self, first_col:str,last_col:str,first_row=2,last_row=0):
      # only area in the dformat A1:Z100 works as argument in conditional format()
      first_col_idx = self.columns.to_list().index(first_col)
      last_col_idx = self.columns.to_list().index(last_col)

      first_col_letter_idx = string.ascii_uppercase[first_col_idx]
      last_col_letter_idx = string.ascii_uppercase[last_col_idx]

      area = first_col_letter_idx+str(first_row)+':'+last_col_letter_idx
      if last_row:
          area += str(last_row)
      else:
          area += str(first_row+len(self)-1)
      
      return area

  
  def set_conditional_frmt(self,conditional_format:dict, for_first_col:str,and_last_col:str,for_first_row=1,and_last_row=0):
      lr = and_last_row if and_last_row else len(self)+1
      area = self.__worksheet_area4(for_first_col,and_last_col,for_first_row,lr)
      self.conditional_frmt[area] = conditional_format


  def make_header_vertical(self,height=120):
      self.header_format['rotation'] = '90'
      self.header_format['height'] = height
      self.header_format['valign'] = 'vcenter'


  def make_header_horizontal(self):
      self.header_format.pop('rotation','not_found')
      self.header_format.pop('height','not_found')
      self.header_format.pop('valign','not_found')


  def __df2excel(self,writer:ExcelWriter,sheet_name:str,**kwargs):
    '''
    Column format specifications must be in self.column2format\n
    Header format specification must be in self.header_format\n
    Tab format specification must be in tab_format

    Parameters
    ----------
    height,width,wrap_text,inch_width\n
    other format parameters are at https://xlsxwriter.readthedocs.io/format.html
    '''
    my_kwargs = {'startrow':1,'header':False,'index':False,'float_format':'%g'}
    my_kwargs.update(kwargs)

    self.to_excel(writer, sheet_name=sheet_name, **my_kwargs)
    assert isinstance(writer.book,xlsxwriter.workbook.Workbook)

    my_worksheet = writer.sheets[sheet_name]
    assert isinstance(my_worksheet,xlsxwriter.workbook.Worksheet)
    # writing header
    header_height = self.header_format.pop('height',15)
    
    format = writer.book.add_format(self.header_format)
    my_worksheet.set_row(0,header_height,format)
    for col_num, value in enumerate(self.columns.values):
      writer.sheets[sheet_name].write(0, col_num, value,format)

    # formating columns
    for idx, column_name in enumerate(self.columns.to_list()):
      try:
        col_fmt_dic = dict(self.column2format[column_name])
        # have to pop format settings unfamiliar to column_format
        width = col_fmt_dic.pop('width', 20)
        wrap = col_fmt_dic.pop('wrap_text', False)
        col_fmt_dic.pop('inch_width', 1)
        column_format = writer.book.add_format(col_fmt_dic)
        if wrap: column_format.set_text_wrap()
        my_worksheet.set_column(idx,idx,width,column_format)
      except KeyError:
        continue

    for area, fmt in self.conditional_frmt.items():
      #_1row,_1col,last_row,last_col = list(area)
      #my_worksheet.conditional_format(_1row,_1col,last_row,last_col,fmt)
      my_worksheet.conditional_format(area, fmt) # only area in format A1:Z100 works here

    try:
      tab_color = self.tab_format['tab_color']
      my_worksheet.set_tab_color(tab_color)
    except KeyError:
      pass


  def df2excel(self,writer:ExcelWriter,sheet_name:str,**kwargs):
      if len(self) > 1000000:
          chunks = [df.from_pd(self[i:i+1000000]) for i in range(0, len(self), 1000000)]
          [d.__df2excel(writer,sheet_name+str(i+1),**kwargs) for i,d in enumerate(chunks)]
      else:
          self.__df2excel(writer,sheet_name,**kwargs)


  def _2excel(self, fpath:str,ws_name='',mode='w'):
      if not ws_name:
          ws_name = self._name_ if self._name_ else 'Sheet1'

      f = ExcelWriter(fpath, engine='xlsxwriter',mode=mode)
      self.df2excel(f,ws_name)
      f.close()


  def clean(self):
      clean_pd = self.dropna(how='all',subset=None)
      clean_pd = clean_pd.drop_duplicates()
      clean_pd = clean_pd.fillna('')
      clean_df = df.from_pd(clean_pd, dfname=self._name_)
      clean_df.copy_format(self)
      return clean_df


  def filter_by(self,values:list, in_column:str):
      copedf = df.from_pd(self[self[in_column].isin(values)],self._name_)
      copedf.copy_format(self)
      return copedf


  def greater_than(self, value:float, in_column:str):
      old_len = len(self)
      new_pd = self[self[in_column] > value]
      removed_rows = old_len - len(new_pd)
      print(f'{removed_rows} rows were removed from {self._name_} because "{in_column}" value was smaller than {value}')
      return df.from_pd(new_pd)
  
  
  def smaller_than(self, value:float, in_column:str):
      return df.from_pd(self[self[in_column] < value])


  def drop_empty_columns(self, max_abs=0.00000000000000001, subset=list()):
      my_columns = subset if subset else self.columns
      my_numeric_columns = subset if subset else [x for x in my_columns if is_numeric_dtype(self[x])]
      no_value_cols = list()
      [no_value_cols.append(col) for col in my_numeric_columns if abs(self[col].max()) < max_abs]
      no_empty_cols = df.from_pd(self.drop(columns = no_value_cols),dfname=self._name_)
      no_empty_cols.copy_format(self)
      print(f'{len(no_value_cols)} columns were dropped because they have all values = 0')
      if len(no_value_cols) < 20:
          print('Dropped columns:')
          [print(c) for c in no_value_cols]
      return no_empty_cols
      

  def table_layout(self, table_witdh=7.5):
      """
      Returns {col_idx:width} for default table_witdh
      """
      length_averages = dict()
      columns = list(self.columns)
      rownum = float(len(self.index))
      col_idx = 0
      for col in columns:
          col_values = list(self[col])+[col]
          col_values = list(map(str,col_values))
          split_col_values = list()
          for v in col_values:
              sentences = v.split('\n')
              split_col_values += sentences
          col_lengths = list(map(len, split_col_values))
          length_averages[col_idx] = float(sum(col_lengths))/rownum
          col_idx +=1
      
      scale_factor = table_witdh/sum(length_averages.values())
      normalized_layout = {k:v*scale_factor for k,v in length_averages.items()}

      # columns cannot be too narrow
      #min_width = 0.65
      [normalized_layout.update({i:MIN_COLUMN_WIDTH}) for i,w in normalized_layout.items() if w < MIN_COLUMN_WIDTH]

      # trying to fit the table into the page by narrowing very wide columns
      new_table_width = sum(list(normalized_layout.values()))
      width_excess = new_table_width-table_witdh
      for col_idx,col_width in normalized_layout.items():
          if col_width > 3:
              new_width = col_width - width_excess
              if new_width > 2:
                  normalized_layout[col_idx]= new_width
              break

      new_table_width = sum(list(normalized_layout.values()))
      width_excess = new_table_width-table_witdh
      if width_excess > 0:
          #distributing width excess across all columns wider than min_col_width
          wide_col_idxes = list()
          sum_width2trim = 0.0
          for idx,width in normalized_layout.items():
              if width > MIN_COLUMN_WIDTH:
                  wide_col_idxes.append(idx)
                  sum_width2trim =  sum_width2trim + width

          scale_factor = sum_width2trim/(sum_width2trim+width_excess)
      
          for col_idx,col_width in normalized_layout.items():
              if col_idx in wide_col_idxes:
                  normalized_layout[col_idx] = col_width*scale_factor
      
      #table_width = sum(list(normalized_layout.values()))
      # assert table_width <= 6.8
      return normalized_layout


  def sort_columns_by_list(self,only_columns:list):
      my_columns = [c for c in only_columns if c in self.columns]
      new_df = df.from_pd(self[my_columns])
      new_df.copy_format(self)
      return new_df


  def clean4doc(self,max_row=int(),only_columns=[],ref_limit=dict(), as_str=True):    
      clean_df = self.sort_columns_by_list(only_columns) if only_columns else self.dfcopy()
      clean_df.dropna(how='all',subset=None,inplace=True)
      clean_df.drop_duplicates(inplace=True)
      clean_df.fillna('', inplace=True)

      if max_row: clean_df = df.from_pd(clean_df.head(max_row))

      if ref_limit:
          for col_name, cutoff in ref_limit.items():
              clean_df = df(clean_df.loc[clean_df[col_name] >= cutoff])

      if as_str:
          for col in clean_df.columns:
              if clean_df[col].dtype in ['float','float64']:
                  clean_df[col] = clean_df[col].map(lambda x: '%2.2f' % x)
              clean_df[col] = clean_df[col].astype(str)
      return clean_df


  @staticmethod
  def read_clean(*args,**kwargs):
      """
      Input
      -----
      file name must be in args[0]
      'only_columns' = [col_names]
      'ref_limit' = [col_name:reflimit]
      'as_str' = True - formats floats to string as %2.2f'
      max_row = 0 - read only first max_row rows
      other *args,**kwargs as in pandas.read_excel, pandas.read_excel
      """
      only_columns = kwargs.pop('only_columns',[])
      ref_limit = kwargs.pop('ref_limit',dict())
      as_str = kwargs.pop('as_str',True)
      max_row = kwargs.pop('max_row',0)
      try:
          table_df = df.read(*args,**kwargs)
          if table_df.empty:
              print('worksheet %s in %s is empty' % (table_df._name_,args[0]))
              return df()
          
          if only_columns:
              if isinstance(only_columns[0],int):
                  select_columns = [v for i,v in enumerate(table_df.columns.to_list()) if i in only_columns]
              else:
                  select_columns = only_columns
          else:
              select_columns = []

          return table_df.clean4doc(max_row,select_columns,ref_limit,as_str) 
      except FileNotFoundError: 
          raise FileNotFoundError


  @staticmethod
  def formula2hyperlink(formula:str):
      '''
      extracts hyperlink from HYPERLINK formula in Excel
      '''
      values = formula[formula.find('(')+1:-2]
      url, txt = values.split(',')
      return url.strip(' "'), txt.strip(' "')


  def remove_rows_by(self, values:list, in_column:str):
      clean_df = df.from_pd(self[~self[in_column].isin(values)])
      clean_df.__copy_attrs(self)
      return clean_df


  def is_numeric(self,column:str):
      return is_numeric_dtype(self[column])


  def to_dict(self,key_col:str,values_col:str):
      return dict(list(zip(getattr(self,key_col),getattr(self,values_col))))


  def not_nulls(self,columns4count:list,write2column='Row count'):
      '''
      Adds
      ----
      Column named "write2column" with count of empty "columns4count" for every row
      '''
      self[write2column] = self[columns4count].isna().sum(axis=1)
      self[write2column] = self[write2column].apply(lambda x: len(columns4count) - x)


  def add_values(self,from_column:str,in_df:'df',map_by_my_col:str,to_my_col='',map2col='',how2replace='false'):
      '''
      Input
      -----
      replace = 'false' - no replacement\n
      replace = 'true' - replace existing values\n
      replace = 'merge' - adds values to existing values after ";"

      Returns
      -------
      df with new "to_my_col"\n
      if "to_my_col" is not specified the column named "from_column" will be created in retruned df\n
      if "map2col" is not specified mapping of "in_df" values uses column with name "map_by_my_col" that must be present "in_df" 
  
      '''
      in_df_map_column = map2col if map2col else map_by_my_col
      map_dict = in_df.to_dict(in_df_map_column,from_column)
      copy2column =  to_my_col if to_my_col else from_column
      def __my_value(x):
          try:
              new_value = map_dict[x[map_by_my_col]]
          except KeyError:
              new_value = ''

          exist_value = '' if pd.isna(x[copy2column]) else str(x[copy2column])

          if how2replace == 'true':
              return new_value
          elif how2replace == 'merge':
              if exist_value:
                  return ';'.join({exist_value,new_value}) if new_value else exist_value
              else:
                  return new_value if new_value else np.nan
          else:
              # if how2replace='false' - do not replace
              if exist_value:
                  return exist_value
              else:
                  return new_value if new_value else np.nan

      copy_df = self.dfcopy()
      if copy2column not in copy_df.columns: copy_df[copy2column] = np.nan
      copy_df[copy2column] = copy_df.apply(__my_value, axis=1)
      return copy_df


  def move_cols(self,col2pos:dict[str,int]):
      '''
      input:
          col2pos = {colulmn_name : position}
      '''
      my_columns = [c for c in self.columns.to_list() if c not in col2pos]
      sorted_col2pos = dict(sorted(col2pos.items(), key=lambda item: item[1]))
      [my_columns.insert(pos,c) for c,pos in sorted_col2pos.items() if c in self.columns]
      return self.dfcopy(my_columns)
          

  def l2norm(self,columns:list|dict=[]):
      '''
      Input
      -----
      columns = {column2normalize:column_with_normalization}
      '''
      my_cols = columns if columns else self.columns
      copy_df = self.dfcopy()
      for col in my_cols:
        if is_numeric_dtype(copy_df[col]):
          vec = copy_df[[col]].to_numpy(float)
          veclen =  np.sqrt(np.sum(vec**2))
          new_col = my_cols[col] if isinstance(my_cols,dict) else col 
          copy_df[new_col] = vec / veclen if veclen > 0.0 else 0.0
      
      return copy_df
  

  def minmax_norm(self,columns:list):
      '''
      Input
      -----
      columns = {column2normalize:column_with_normalization}
      ''' 
      my_cols = columns if columns else self.columns
      copy_df = self.dfcopy()
      for col in my_cols:
        if is_numeric_dtype(copy_df[col]):
          vec = copy_df[[col]].to_numpy(float)
          vec_clean = vec[~np.isnan(vec)]
          vec_min = vec_clean.min()
          vec_max = vec_clean.max()
          copy_df[col] = (vec - vec_min) / (vec_max - vec_min)
      return copy_df


  def split(self,num_parts:int):
      split_dataframes = np.array_split(self, num_parts)
      splits = list()
      for d in split_dataframes:
          part = df.from_pd(pd.DataFrame(d,columns=self.columns.to_list()))
          part.__copy_attrs(self)
          splits.append(part)
      
      return splits


  def sort_columns(self,column_aggregate_func,ascending=False):
      col_aggregate_values = self.apply(column_aggregate_func) # Calculate the sum of each column
      assert(isinstance(col_aggregate_values,pd.Series))
      new_order = col_aggregate_values.sort_values(ascending=ascending).index # Sort the column sums
      sorted_self = df.from_pd(self[new_order],dfname=self._name_)
      sorted_self.copy_format(self)
      return sorted_self # Rearrange the columns based on the sorted sums
  
  

  @staticmethod
  def calculate_pvalues(scores:pd.Series,skip1strow=False):
    if skip1strow:
      """Calculates exponential p-values, skipping the first row."""
      if scores.empty or len(scores) <= 1: #handle empty, or single row series.
          return pd.Series([])
      scores_to_fit = np.array(scores.iloc[1:])
      p_values = [(scores_to_fit >= score).mean() for score in scores_to_fit]
      full_p_values = pd.Series(index = scores.index)
      full_p_values.iloc[1:] = p_values
      return full_p_values
    else:
      """Calculates empirical p-values for a list of scores."""
      scores = np.array(scores)
      p_values = [(scores >= score).mean() for score in scores]
      return p_values
  

  @staticmethod
  def calculate_expo_pvalues(scores:pd.Series,skip1strow=False):
    if skip1strow:
      """Calculates exponential p-values, skipping the first row."""
      if scores.empty or len(scores) <= 1: #handle empty, or single row series.
          return pd.Series([])

      scores_to_fit = scores.iloc[1:]  # Skip the first row
      scores_filled = scores_to_fit.fillna(0)
      lambda_hat = expon.fit(scores_filled, floc=0)[1]
      if lambda_hat <= 0:
        print(f"Warning: lambda_hat is {lambda_hat}. Returning array of ones.")
        p_values = pd.Series(np.ones_like(scores_filled), index=scores_filled.index)
      else:
        p_values = expon.sf(scores_filled, scale=lambda_hat)
        p_values = pd.Series(p_values, index=scores_filled.index)

      full_p_values = pd.Series(index = scores.index)
      full_p_values.iloc[1:] = p_values
      return full_p_values
    else:
      scores_filled = scores.fillna(0)
      lambda_hat = expon.fit(scores_filled, floc=0)[1] # Fit an exponential distribution
      p_values = expon.sf(scores_filled, scale=lambda_hat) # Calculate p-values
      return p_values
  

  def sortrows(self,by:str|list[str],ascending:bool|list[bool]=False,skip_rows=0):
    # re-writing parent pd.DataFrame function is a BAD idea. Use other names
    if skip_rows:
      if len(self) > skip_rows:
        top_rows = self.iloc[:skip_rows]
        remaining_rows = self.iloc[skip_rows:].sort_values(by=by, ascending=ascending,inplace=False)
        sorted_df = df.from_pd(pd.concat([top_rows, remaining_rows], ignore_index=True),dfname=self._name_)
        sorted_df.copy_format(self)
        return sorted_df
      else:
        return self
    else:
      sorted_df = pd.DataFrame(self.sort_values(by,ascending=ascending,inplace=False))
      sorted_df = df.from_pd(sorted_df,self._name_)
      sorted_df.copy_format(self)
    return sorted_df
    

  @staticmethod
  def info_df()->"df":
    rows = [['Number of worksheets in this file:','=INFO("numfile")']]
    return df.from_rows(rows,['Info','Counts'],dfname='info')
